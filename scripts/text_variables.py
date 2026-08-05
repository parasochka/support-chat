#!/usr/bin/env python3
"""The TEXT-VARIABLES reference — every placeholder and text slot of the product.

Sibling of the API reference (scripts/build_api_reference.py), deliberately the
OTHER half of it: that table carries only what crosses a system boundary, this
one carries what lives INSIDE the product and is edited as text — the KB
`{placeholder}` registry, the prompt-variable registries (support + Telegram),
the user-facing copy registry, the placeholders that live inside that copy, the
control sentinels the model emits, the large editable text blocks (KB texts,
site map, escalation keywords) and the player-profile fields the prompt renders.
Audience: whoever fills a new brand in — content owner, ops, integrator.

ONE row = one text unit: a variable, a copy key, a sentinel or a text block.
Every column is a filter, so the same sheet answers "what do I have to replace
per brand", "what is edited where in the admin", "what must stay English".

The rows are DERIVED FROM THE LIVE REGISTRIES (imported, not transcribed), so
the page cannot drift from the code: db._DEFAULT_KB_VARIABLES,
prompts.PROMPT_VARIABLES, prompts.RETENTION_PROMPT_VARIABLES,
translations.KEYS/DEFAULTS and prompts._CONTEXT_FIELDS. Only the hand-written
part (the sentinels, the text blocks, the per-group notes) is prose, and the
sentinel list is checked against the strip-regexes in prompts.py at build time.

Not a script: it is rendered by scripts/build_api_reference.py (one command
builds the page and the second sheet of the same workbook, so page and workbook
can never drift apart):

  frontend/api-reference.xlsx            — sheet "Переменные текста"
  frontend/integration-variables.html    — live page (/integration-variables)
"""
from __future__ import annotations

import html
import importlib.util
import os
import re
import sys
from typing import Any

SHEET_TITLE = "Переменные текста"

# ---------------------------------------------------------------------------
# Columns. Order matters: it is the column order in both renderings.
# ---------------------------------------------------------------------------
COLUMNS: list[tuple[str, int]] = [
    ("Группа", 26),
    ("Ключ / как пишется в тексте", 32),
    ("Где используется", 28),
    ("Где редактируется", 28),
    ("Скоуп", 18),
    ("Язык значения", 18),
    ("Значение по умолчанию / пример", 54),
    ("Описание", 62),
]

MONO_COLS = {1, 6}
FILTER_COLS = [0, 2, 3, 4, 5]

# --- группы ---------------------------------------------------------------
G_KB = "Переменные базы знаний"
G_TODO = "Незаполненные значения бренда"
G_PROMPT = "Промпт-переменные (поддержка)"
G_RPROMPT = "Промпт-переменные (Telegram)"
G_COPY_W = "Тексты: виджет"
G_COPY_S = "Тексты: ответы сервера"
G_COPY_R = "Тексты: Telegram-бот"
G_INLINE = "Плейсхолдеры внутри текстов"
G_TAG = "Управляющие теги модели"
G_BLOCK = "Крупные тексты и справочники"
G_CTX = "Поля профиля игрока в промпте"

# --- где используется ------------------------------------------------------
U_KB = "Тексты базы знаний"
U_CORE = "Системный промпт (Layer 1)"
U_L3 = "Промпт, данные хода (Layer 3)"
U_BOTH = "Оба промпта (чат + Telegram)"
U_WIDGET = "Виджет на сайте"
U_SERVER = "Ответы сервера в чате"
U_TG = "Telegram-бот"
U_OUT = "Ответ модели (служебный тег)"

# --- где редактируется -----------------------------------------------------
E_KBVAR = "Knowledge base → Variables"
E_KB = "Knowledge base → тексты тем"
E_PV = "Prompt → Prompt variables"
E_RPV = "Retention → Prompt variables"
E_TR = "Translations"
E_SITE = "Common → Site map"
E_ESC = "Common → Escalation keywords"
E_PROFILE = "Common → Test profile"
E_RKB = "Retention → Knowledge base"
E_SCEN = "Retention → Scenarios"
E_CODE = "код prompts.py (редеплой)"
E_CRM = "приходит из CRM казино"

# --- скоуп ------------------------------------------------------------------
S_PRODUCT = "продукт"
S_BOTH = "глобально + продукт"
S_DEPLOY = "деплой (общий)"

# --- язык -------------------------------------------------------------------
L_EN = "английский (проверяется)"
L_MULTI = "мультиязычно"
L_NONE = "—"


def row(group: str, key: str = "", syntax: str = "", used: str = "",
        edited: str = "", scope: str = "", lang: str = "", default: str = "",
        desc: str = "") -> list[str]:
    """One text unit. `syntax` (how the key is written inside a text) wins over
    `key` in the rendered column — for a variable the two say the same thing
    twice, and the braces are the half that matters when you are writing copy."""
    return [group, syntax or key, used, edited, scope, lang, default, desc]


# ---------------------------------------------------------------------------
# Live registries. Imported through the test bootstrap so db.py (which imports
# asyncpg) loads without the real driver — exactly like scripts/check_invariants.py.
# ---------------------------------------------------------------------------
def _load_app_modules() -> tuple[Any, Any, Any]:
    os.environ.setdefault("SUPPORT_CHAT_TEST_MODE", "1")
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if root not in sys.path:
        sys.path.insert(0, root)
    spec = importlib.util.spec_from_file_location(
        "_conftest_bootstrap", os.path.join(root, "tests", "conftest.py"))
    if spec and spec.loader:
        spec.loader.exec_module(importlib.util.module_from_spec(spec))

    from app.ai import prompts
    from app.core import db
    from app.i18n import translations

    return prompts, db, translations


def _shorten(text: str, limit: int = 320) -> str:
    text = " ".join(str(text or "").split())
    return text if len(text) <= limit else text[:limit - 1].rstrip() + "…"


# ---------------------------------------------------------------------------
# Hand-written rows: the control sentinels the model emits inside its answer.
# `attr` is the strip-regex in prompts.py — checked at build time so a renamed
# or removed sentinel breaks the build instead of silently outliving the code.
# ---------------------------------------------------------------------------
_SENTINELS: list[tuple[str, str, str, str, str]] = [
    ("_ESCALATE_TAG_RE", "[[ESCALATE]]", "[[ESCALATE]]", U_OUT,
     "Передача вопроса человеку. Бэкенд вырезает тег, закрывает сессию и "
     "показывает карточку контакта (или диплинк в Telegram-бота)."),
    ("_TOPIC_TAG_RE", "[[TOPIC:slug]]", "[[TOPIC:withdrawals]]", U_OUT,
     "Вопрос относится к другой теме. Бэкенд подавляет текст ответа "
     "(он сгенерирован без нужной базы знаний), переключает тему и "
     "переспрашивает исходный вопрос."),
    ("_SUGGEST_TAG_RE", "[[SUGGEST: a | b]]",
     "[[SUGGEST: Какие лимиты на вывод? | Сколько идут выплаты?]]", U_OUT,
     "До двух уточняющих вопросов от лица игрока — кнопки в один тап. "
     "Опция «Проблема решена.» добавляется сервером, а не моделью."),
    ("_RESOLVED_TAG_RE", "[[RESOLVED]]", "[[RESOLVED]]", U_OUT,
     "Вопрос закрыт — виджет показывает зелёную кнопку завершения чата."),
    ("_LANG_TAG_RE", "[[LANG:xx]]", "[[LANG:ru]]", U_OUT,
     "Язык, на котором модель ответила. Бэкенд запоминает его на сессии, "
     "интерфейс переключается следом за игроком."),
    ("_PHOTO_TAG_RE", "[[PHOTO:id]]", "[[PHOTO:42]]", U_OUT,
     "Только Telegram: отправить фото/видео с этим id из списка кандидатов, "
     "показанного модели. Id проверяется по разрешённому набору."),
    ("_LINK_TAG_RE", "[[LINK:url]]", "[[LINK:https://brand.com/cashier]]", U_OUT,
     "Только Telegram: подставить кнопку на страницу из карты сайта. "
     "Адрес вне карты сайта отбрасывается."),
    ("_STAGE_UP_TAG_RE", "[[STAGE_UP]]", "[[STAGE_UP]]", U_OUT,
     "Только Telegram: подсказка, что игрок готов к следующей ступени "
     "откровенности. Решение принимает бэкенд, а не модель."),
    ("_HANDOFF_TAG_RE", "[[HANDOFF]]", "[[HANDOFF]]", U_OUT,
     "Только Telegram: всплыл вопрос поддержки/жалоба — бот уводит игрока "
     "к менеджеру или в чат поддержки на сайте."),
]

# Hand-written rows: the big editable text blocks. Not placeholders, but the
# same question ("where does this text live") — the page is the map.
_BLOCKS: list[tuple[str, str, str, str, str, str, str]] = [
    ("Тексты тем базы знаний", U_KB, E_KB, S_PRODUCT, L_EN,
     "JSON-документ вопрос-ответ на тему",
     "Содержимое ответов (Layer 2): грузится только по выбранной теме. "
     "Внутри можно ставить {переменные} из реестра выше."),
    ("Названия тем", U_WIDGET, E_TR, S_PRODUCT, L_MULTI,
     "Депозиты / Выводы / Бонусы …",
     "Заголовки кнопок выбора темы — по одному на язык. Каноническое "
     "английское название темы редактируется вместе с самой темой."),
    ("База знаний Telegram-бота", U_TG, E_RKB, S_PRODUCT, L_EN,
     "один документ на продукт",
     "Отдельный от поддержки KB-документ ретеншен-персоны. Те же "
     "{переменные} реестра подставляются и в него."),
    ("Карта сайта (title / url / purpose)", U_BOTH, E_SITE, S_BOTH, L_EN,
     '{"title":"Cashier","url":"https://brand.com/cashier",'
     '"purpose":"deposits and withdrawals"}',
     "Единственный список страниц, на которые обоим ботам разрешено ставить "
     "ссылки. Пустой список = блока в промпте нет."),
    ("Ключевые слова эскалации", "Сообщение игрока (до вызова модели)", E_ESC,
     S_BOTH, L_MULTI, "high_risk_keywords · human_request_keywords",
     "Два списка основ слов: жалоба/мошенничество/юридическая угроза и "
     "явная просьба человека. Сканируют сообщение игрока, поэтому "
     "мультиязычны и на английский не проверяются."),
    ("Тест-профиль игрока", U_L3, E_PROFILE, S_BOTH, L_NONE,
     '{"full_name":"Andrey Petrov","vip_level":"Gold"}',
     "Профиль, который подставляется в предпросмотр промпта, чтобы видеть "
     "персонализацию без живой сессии."),
    ("Шаблоны сценариев ретеншена", U_TG, E_SCEN, S_PRODUCT, L_EN,
     "persona_brief (по умолчанию) | verbatim",
     "Шаг journey задаёт НАМЕРЕНИЕ, текст пишет персона. Точный текст "
     "используется только при явном флаге verbatim."),
    ("Формулировки промпта (SYSTEM_CORE, директивы, "
     "запрещённые темы, теги качества)", U_CORE, E_CODE, S_DEPLOY, L_EN,
     "prompts.py",
     "Сами формулировки промпта — единственный источник правды и НЕ "
     "редактируются из админки: бренд уникализируется переменными выше."),
]

_CTX_DESC = {
    "id": "Идентификатор игрока в казино.",
    "full_name": "Полное имя. Из него берётся ИМЯ для приветствия в первом "
                 "ответе; без имени приветствия нет вообще.",
    "email": "Почта игрока.",
    "activation_status": "Статус активации/верификации.",
    "country": "Страна.",
    "balance": "Баланс строкой, вместе с валютой.",
    "vip_level": "VIP-класс.",
    "registration_date": "Дата регистрации.",
}

_INLINE_DESC = {
    "topic": "Название темы, на которую переключается чат.",
    "persona": "Имя персоны — подставляется из промпт-переменных.",
    "name": "Имя игрока из профиля.",
    "manager": "Ссылка или контакт назначенного менеджера.",
    "detail": "Деталь события (сумма, уровень, тип бонуса); может быть пустой.",
}


def _build_rows() -> list[list[str]]:
    prompts, db, translations = _load_app_modules()
    rows: list[list[str]] = []

    # --- KB variables ------------------------------------------------------
    for key, desc, value in db._DEFAULT_KB_VARIABLES:
        rows.append(row(
            G_KB, key, "{%s}" % key, U_KB, E_KBVAR, S_PRODUCT, L_EN,
            _shorten(value), desc))

    # --- {{PLACEHOLDER}} values that MUST be replaced per brand -------------
    todo: dict[str, list[str]] = {}
    for key, _desc, value in db._DEFAULT_KB_VARIABLES:
        for token in re.findall(r"\{\{([A-Z0-9_]+)\}\}", value):
            todo.setdefault(token, []).append(key)
    for token, keys in sorted(todo.items()):
        rows.append(row(
            G_TODO, token, "{{%s}}" % token, U_KB, E_KBVAR, S_PRODUCT, L_EN,
            "переменные: " + ", ".join(keys),
            "Заглушка в значении по умолчанию: реестр засевается всем "
            "продуктам, поэтому имя бренда, ссылка или лицензия проставляются "
            "владельцем. Пока стоит заглушка, модель отдаёт её игроку как есть."))

    # --- prompt variables (support) ----------------------------------------
    for key, desc, default in prompts.PROMPT_VARIABLES:
        rows.append(row(
            G_PROMPT, key, "{%s}" % key, U_CORE, E_PV, S_BOTH, L_EN,
            _shorten(default), desc))

    # --- prompt variables (retention / Telegram) ---------------------------
    for key, desc, default, renders_as in prompts.RETENTION_PROMPT_VARIABLES:
        note = (f" В шаблонах ретеншена заполняет базовый плейсхолдер "
                f"{{{renders_as}}} — связь по РЕНДЕРУ, не по значению: "
                f"правка в поддержке в бота не протекает."
                if renders_as else
                f" В шаблонах ретеншена пишется как {{{key}}}.")
        rows.append(row(
            G_RPROMPT, key, "", U_CORE, E_RPV,
            S_BOTH, L_EN, _shorten(default), desc + note))

    # --- user-facing copy registry -----------------------------------------
    scope_group = {"widget": (G_COPY_W, U_WIDGET),
                   "server": (G_COPY_S, U_SERVER),
                   "retention": (G_COPY_R, U_TG)}
    english = translations.DEFAULTS.get("en", {})
    for key, scope, desc in translations.KEYS:
        group, used = scope_group.get(scope, (G_COPY_S, U_SERVER))
        rows.append(row(
            group, key, "", used, E_TR, S_BOTH, L_MULTI,
            _shorten(english.get(key, "")), desc))

    # --- placeholders that live INSIDE that copy ---------------------------
    inline: dict[str, list[str]] = {}
    for key, text in english.items():
        for token in re.findall(r"\{([a-z_]+)\}", text):
            inline.setdefault(token, []).append(key)
    for token, keys in sorted(inline.items()):
        rows.append(row(
            G_INLINE, token, "{%s}" % token, U_WIDGET + " / " + U_TG, E_TR,
            S_BOTH, L_NONE, "ключи: " + ", ".join(sorted(set(keys))),
            _INLINE_DESC.get(token, "Подстановка внутри строки копирайта.") +
            " Плейсхолдер обязан сохраниться при переводе — иначе подстановка "
            "потеряется."))

    # --- control sentinels --------------------------------------------------
    for attr, syntax, example, used, desc in _SENTINELS:
        if not hasattr(prompts, attr):
            raise SystemExit(
                f"text_variables: sentinel {syntax} has no {attr} in prompts.py "
                "— update the registry in scripts/text_variables.py")
        rows.append(row(
            G_TAG, syntax.strip("[]").split(":")[0], syntax, used, E_CODE,
            S_DEPLOY, L_NONE, example,
            desc + " Тег вырезается из ответа и игроку не показывается."))

    # --- large editable text blocks ----------------------------------------
    for name, used, edited, scope, lang, default, desc in _BLOCKS:
        rows.append(row(G_BLOCK, name, "", used, edited, scope, lang,
                        default, desc))

    # --- player-profile fields the prompt renders --------------------------
    for field in prompts._CONTEXT_FIELDS:
        rows.append(row(
            G_CTX, field, "", U_L3, E_CRM, S_PRODUCT, L_NONE, "",
            _CTX_DESC.get(field, "") +
            " Белый список: всё остальное из профиля до модели не доходит. "
            "Контракт передачи — на /integration-reference."))

    return rows


ROWS: list[list[str]] = _build_rows()

SUBTITLE = ("Все переменные и куски текста, из которых собирается общение с "
            "игроком: подстановки базы знаний, переменные промпта, копирайт "
            "интерфейса и служебные теги. Одна строка — одна текстовая "
            "единица, все колонки фильтруются.")


# ---------------------------------------------------------------------------
# Excel sheet (второй лист того же файла)
# ---------------------------------------------------------------------------
def add_sheet(wb, style: dict) -> None:
    """Append the variables sheet to an already-built workbook."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(SHEET_TITLE)
    ws["A1"] = "Support Chat — переменные и тексты продукта"
    ws["A1"].font = style["title"]
    ws.row_dimensions[1].height = 22
    ws["A2"] = SUBTITLE
    ws["A2"].font = style["note"]

    head = 4
    for i, (name, _w) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=head, column=i, value=name)
        c.font = style["head_font"]
        c.fill = style["head_fill"]
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = style["border"]
    ws.row_dimensions[head].height = 30

    r = head
    group_fill = style["group_fill"]
    seen: set[str] = set()
    for data in ROWS:
        r += 1
        first_of_group = data[0] not in seen
        seen.add(data[0])
        for i, val in enumerate(data, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.font = style["mono"] if (i - 1) in MONO_COLS else style["body"]
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = style["border"]
            if first_of_group:
                c.fill = group_fill

    for i, (_name, w) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=head + 1, column=1)
    ws.auto_filter.ref = f"A{head}:{get_column_letter(len(COLUMNS))}{r}"

    r += 2
    ws.cell(row=r, column=1,
            value="Лист генерируется из кода: scripts/build_api_reference.py "
                  "(данные — scripts/text_variables.py). Живая версия — "
                  "/integration-variables на самом сервисе."
            ).font = style["note"]


# ---------------------------------------------------------------------------
# Живая страница — та же таблица с фильтрами и поиском
# ---------------------------------------------------------------------------
PAGE = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Support Chat — переменные и тексты продукта</title>
<style>
  :root {{
    --bg:#0f1320; --panel:#1a2032; --panel2:#232c44; --line:#2e3a57;
    --text:#e7ecf5; --muted:#9aa7c2; --accent:#4f8cff; --good:#36c08a; --warn:#e8b349;
  }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; background:var(--bg); color:var(--text);
    font:14px/1.6 -apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif; }}
  .wrap {{ max-width:1600px; margin:0 auto; padding:28px 20px 70px; }}
  h1 {{ font-size:25px; margin:0 0 6px; }}
  .muted {{ color:var(--muted); }}
  a {{ color:var(--accent); }}
  code {{ background:var(--panel2); border:1px solid var(--line); border-radius:5px;
    padding:1px 6px; font:12.5px/1.5 ui-monospace,Menlo,Consolas,monospace; }}
  .dl {{ display:inline-block; background:rgba(54,192,138,.14); border:1px solid var(--good);
    color:var(--good); border-radius:8px; padding:9px 16px; text-decoration:none;
    font-weight:600; margin:10px 0 4px; }}
  .dl:hover {{ background:rgba(54,192,138,.24); }}
  .filters {{ background:var(--panel); border:1px solid var(--line); border-radius:10px;
    padding:14px 16px; margin:18px 0; display:flex; flex-wrap:wrap; gap:10px 14px;
    align-items:flex-end; position:sticky; top:0; z-index:5; }}
  .filters label {{ display:block; font-size:11px; text-transform:uppercase;
    letter-spacing:.04em; color:var(--muted); margin-bottom:4px; }}
  select, input[type=search] {{ background:var(--panel2); color:var(--text);
    border:1px solid var(--line); border-radius:7px; padding:7px 10px; font-size:13px;
    min-width:150px; }}
  input[type=search] {{ min-width:240px; }}
  .count {{ margin-left:auto; color:var(--muted); font-size:13px; white-space:nowrap; }}
  .reset {{ background:none; border:1px solid var(--line); color:var(--muted);
    border-radius:7px; padding:7px 12px; font-size:13px; cursor:pointer; }}
  .reset:hover {{ color:var(--text); border-color:var(--accent); }}
  /* The table scrolls in its OWN pane: that is what lets the header row stay
     pinned (a sticky th only sticks inside its scroll container) while the
     filter bar keeps the top of the page. --fh is the live height of that bar,
     set from JS — it wraps to two lines on narrow viewports. */
  .tablewrap {{ overflow:auto; border:1px solid var(--line); border-radius:10px;
    max-height:calc(100vh - var(--fh,150px) - 24px); }}
  table {{ width:100%; border-collapse:collapse; font-size:13px; }}
  th, td {{ text-align:left; padding:8px 10px; border-bottom:1px solid var(--line);
    vertical-align:top; }}
  th {{ color:var(--muted); font-size:11px; text-transform:uppercase; letter-spacing:.04em;
    background:var(--panel); position:sticky; top:0; z-index:2; white-space:nowrap; }}
  tr.first td {{ border-top:2px solid var(--line); }}
  /* A long KB value would otherwise eat the width and push the description off
     screen — the description is the half that answers "what is this". */
  td:nth-child(7) {{ max-width:380px; }}
  td:nth-child(8) {{ min-width:300px; }}
  td.mono {{ font:12.5px/1.5 ui-monospace,Menlo,Consolas,monospace; color:#cbd5e1; }}
  td.key {{ font-weight:600; color:#e7ecf5; }}
  .note {{ background:rgba(79,140,255,.08); border:1px solid var(--line);
    border-left:3px solid var(--accent); border-radius:8px; padding:10px 14px; margin:14px 0; }}
  .legend {{ display:grid; gap:8px 22px; grid-template-columns:repeat(auto-fit,minmax(330px,1fr));
    margin:14px 0 4px; padding:0; list-style:none; }}
  .legend li {{ color:var(--muted); }}
  .legend b {{ color:var(--text); }}
  tr.hidden {{ display:none; }}
</style>
</head>
<body>
<div class="wrap">

<h1>Переменные и тексты продукта</h1>
<p class="muted">{subtitle}</p>
<p><a class="dl" href="/static/{xlsx}" download>⤓ Скачать в Excel (лист «{sheet}»)</a></p>

<div class="note"><b>Как это устроено.</b> Формулировки промпта — единственный
источник правды и живут в коде: бренд уникализируется не правкой промпта, а
значениями переменных. Всё, что подставляется в текст, пишется одинаково —
<code>{{ключ}}</code>: и переменные базы знаний, и переменные промпта, и
подстановки внутри строк копирайта. Значения, которые видит МОДЕЛЬ (база знаний,
переменные промпта, карта сайта), обязаны быть на английском — сервис проверяет
это при сохранении и отвечает 400 на кириллицу; языком ответа управляет отдельная
директива, поэтому игрок всё равно получает ответ на своём языке. Значения,
которые игрок видит НАПРЯМУЮ (копирайт интерфейса, ответы сервера, тексты бота),
наоборот, переводятся по языкам.</div>

<ul class="legend">
  <li><b>Скоуп «продукт»</b> — своё значение у каждого казино.</li>
  <li><b>Скоуп «глобально + продукт»</b> — общее значение, которое продукт может
      переопределить.</li>
  <li><b>Скоуп «деплой»</b> — общее для всей инсталляции, меняется только релизом.</li>
  <li><b>Неизвестный <code>{{ключ}}</code> не подставляется</b> — он остаётся в тексте
      как есть, поэтому опечатка видна в предпросмотре промпта, а не «съедает» текст.</li>
</ul>
<p class="muted" style="margin:2px 0 0">Колонка «где редактируется» — это вкладка
админ-панели (например <code>Knowledge base → Variables</code>); строки с пометкой
<code>код prompts.py</code> меняются только релизом. Значения и описания взяты из рабочих
реестров сервиса как есть, поэтому у модельных переменных они английские.</p>

<div class="filters">{filters}
  <button class="reset" type="button" id="reset">Сбросить</button>
  <span class="count" id="count"></span>
</div>

<div class="tablewrap">
<table id="t">
<thead><tr>{head}</tr></thead>
<tbody>
{body}
</tbody>
</table>
</div>

<p class="muted" style="margin-top:32px">Страница и Excel-файл генерируются из кода одним
скриптом (<code>scripts/build_api_reference.py</code>) прямо из рабочих реестров сервиса,
поэтому не расходятся ни между собой, ни с поведением бота.
Контракты, которые пересекают границу систем (эндпойнты, поля на проводе, события), — на
соседней странице <a href="/integration-reference">справочника API</a>.
Смежные гайды: <a href="/integration">обзор и архитектура</a> ·
<a href="/integration-widget">виджет</a> ·
<a href="/integration-data">данные игрока</a> ·
<a href="/integration-chat-api">Chat API и свой UI</a> ·
<a href="/integration-telegram">retention-бот в Telegram</a> ·
<a href="/integration-admin">интеграция с общей админкой</a>.</p>

</div>
<script>
(function () {{
  var table = document.getElementById('t');
  var rows = Array.prototype.slice.call(table.tBodies[0].rows);
  var selects = Array.prototype.slice.call(document.querySelectorAll('select[data-col]'));
  var search = document.getElementById('q');
  var count = document.getElementById('count');
  var filters = document.querySelector('.filters');

  function measure() {{
    document.documentElement.style.setProperty(
      '--fh', filters.getBoundingClientRect().height + 'px');
  }}
  window.addEventListener('resize', measure);
  measure();

  function apply() {{
    var term = (search.value || '').trim().toLowerCase();
    var shown = 0;
    rows.forEach(function (tr) {{
      var ok = selects.every(function (sel) {{
        if (!sel.value) return true;
        return tr.cells[+sel.dataset.col].textContent.trim() === sel.value;
      }});
      if (ok && term) ok = tr.textContent.toLowerCase().indexOf(term) !== -1;
      tr.classList.toggle('hidden', !ok);
      if (ok) shown++;
    }});
    count.textContent = shown + ' из ' + rows.length + ' строк';
  }}

  selects.forEach(function (s) {{ s.addEventListener('change', apply); }});
  search.addEventListener('input', apply);
  document.getElementById('reset').addEventListener('click', function () {{
    selects.forEach(function (s) {{ s.value = ''; }});
    search.value = '';
    apply();
  }});
  apply();
}})();
</script>
</body>
</html>
"""


def build_html(path: str, xlsx_name: str) -> None:
    parts = []
    for col in FILTER_COLS:
        name = COLUMNS[col][0]
        values = sorted({r[col] for r in ROWS if r[col]})
        opts = "".join(f'<option value="{html.escape(v)}">{html.escape(v)}</option>'
                       for v in values)
        parts.append(
            f'<div><label for="f{col}">{html.escape(name)}</label>'
            f'<select id="f{col}" data-col="{col}">'
            f'<option value="">— все —</option>{opts}</select></div>')
    parts.append('<div><label for="q">Поиск по таблице</label>'
                 '<input type="search" id="q" placeholder="ключ, текст, тема…">'
                 "</div>")

    head = "".join(f"<th>{html.escape(n)}</th>" for n, _w in COLUMNS)

    body = []
    seen: set[str] = set()
    for data in ROWS:
        cls = ' class="first"' if data[0] not in seen else ""
        seen.add(data[0])
        cells = []
        for i, val in enumerate(data):
            text = html.escape(val or "").replace("\n", "<br>")
            if i == 1:
                cells.append(f'<td class="mono key">{text}</td>')
            elif i in MONO_COLS:
                cells.append(f'<td class="mono">{text}</td>')
            else:
                cells.append(f"<td>{text}</td>")
        body.append(f"<tr{cls}>" + "".join(cells) + "</tr>")

    with open(path, "w", encoding="utf-8") as fh:
        fh.write(PAGE.format(subtitle=html.escape(SUBTITLE), xlsx=xlsx_name,
                             sheet=html.escape(SHEET_TITLE),
                             filters="\n".join(parts), head=head,
                             body="\n".join(body)))
