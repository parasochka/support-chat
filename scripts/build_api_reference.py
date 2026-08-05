#!/usr/bin/env python3
"""Generate the integration references — flat tables, page + workbook.

The audience is an INTEGRATOR: another team's programmer who has to make their
system talk to this one. So the table carries only what crosses a system
boundary — endpoints, the fields on the wire, their types and formats, who
calls whom, authorization, limits and error codes. Product-internal knobs (env
vars, hot settings, prompt/KB variables, the copy registry, the DB schema) are
deliberately OUT: nobody outside this service edits them, and they would only
dilute the contract surface that has to be matched against other systems.

ONE row = one integration unit: an endpoint, a field on it, a header, or a
canonical event. Every column is a filter, so the same sheet answers "what does
the partner CRM push us", "what needs auth", "what is mandatory", "what do we
call outward" without a second tab.

Both artefacts are built from the SAME rows, so the live page and the
downloadable workbook can never drift apart:

  frontend/api-reference.xlsx              — downloadable workbook
      sheet "API"                          — the table below
      sheet "Переменные текста"            — scripts/text_variables.py
  frontend/integration-reference.html      — live page (/integration-reference)
  frontend/integration-variables.html      — live page (/integration-variables)

The variables half is the deliberate INSIDE of this table: the KB placeholder
registry, prompt variables, the copy registry, the control sentinels and the
editable text blocks — everything product-internal that this sheet leaves out
on purpose. It lives in scripts/text_variables.py (rows derived from the live
registries) and is rendered here so ONE command keeps all four files in step.

Run:  python scripts/build_api_reference.py

Transcribed from app/api/chat.py, app/api/retention.py, app/api/admin.py,
app/api/admin_auth.py, app/core/auth.py, app/ai/prompts.py and
app/retention/player_sync.py. When a public contract changes, edit ROWS below
and re-run — never hand-edit the generated files.
"""
from __future__ import annotations

import html
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import text_variables  # noqa: E402  (sibling module, not a package)

# ---------------------------------------------------------------------------
# Columns. Order matters: it is the column order in both renderings.
# ---------------------------------------------------------------------------
COLUMNS: list[tuple[str, int]] = [
    ("Область", 16),
    ("Юнит (метод + путь)", 34),
    ("Тип строки", 16),
    ("Направление", 14),
    ("Кто вызывает", 20),
    ("Ключ / поле", 26),
    ("Тип / формат", 30),
    ("Обяз.", 11),
    ("Пример значения", 34),
    ("Авторизация", 30),
    ("Лимиты / идемпотентность", 34),
    ("Коды ошибок", 30),
    ("Описание", 62),
]

# Which columns get the monospace face (0-based).
MONO_COLS = {1, 5, 6, 8}
# Columns rendered as dropdown filters on the HTML page (0-based).
FILTER_COLS = [0, 2, 3, 4, 7]

AREA_CHAT = "Chat API"
AREA_PARTNER = "Partner API"
AREA_TG = "Telegram"
AREA_ADMIN = "Admin API"
AREA_OUT = "Исходящие"

T_EP = "Эндпойнт"
T_REQ = "Поле запроса"
T_RESP = "Поле ответа"
T_HDR = "Заголовок"
T_EVT = "Событие"

IN, OUT, BOTH = "Вход", "Выход", "Вход/Выход"

C_SITE = "Сайт / виджет"
C_CLIENT = "Клиент чата"
C_CRM = "CRM казино"
C_TG = "Telegram"
C_ADMIN = "Внешняя админка"
C_US = "Наш сервис"

YES, NO, COND = "да", "нет", "условно"

# Shorthands for the repeated cells.
AUTH_NONE = "нет (публичный)"
AUTH_SESSION = "Bearer <session token> (JWT HS256)"
AUTH_PARTNER = "Bearer <handshake_secret продукта>"
AUTH_ADMIN = "Bearer <admin JWT> ИЛИ Bearer sak_… (сервисный ключ)"


def row(area, unit, kind, direction, caller, key="", fmt="", req="",
        example="", auth="", limits="", errors="", desc="") -> list[str]:
    return [area, unit, kind, direction, caller, key, fmt, req, example,
            auth, limits, errors, desc]


ROWS: list[list[str]] = []
R = ROWS.append

# ===========================================================================
# Chat API — POST /api/chat/session
# ===========================================================================
_SESSION = "POST /api/chat/session"
R(row(AREA_CHAT, _SESSION, T_EP, BOTH, C_SITE,
      auth=AUTH_NONE + " + Cloudflare Turnstile (advisory)",
      limits="Лимит session:{ip}; сессия создаётся лениво — по выбору темы",
      errors="429 rate_limited · 403 bad_widget_key | no_product | turnstile_failed · "
             "401 bad_handshake · 411 length_required / 413 кап тела (настройка "
             "body_max_bytes, по умолч. 64 KiB; действует на ВСЕ ручки)",
      desc="Создание сессии чата: выдаёт токен и каталог тем. "
           "Точка входа для любого клиента (виджет, своё приложение, мобилка)."))
R(row(AREA_CHAT, _SESSION, T_REQ, IN, C_SITE, "widget_key", "string (wk_…)", COND,
      "wk_7f3a…", desc="Публичный идентификатор продукта (казино). "
      "Отсутствует → используется продукт по умолчанию."))
R(row(AREA_CHAT, _SESSION, T_REQ, IN, C_SITE, "consumer", "string ('web')", NO, "web",
      desc="Метка фасада. Любое иное значение схлопывается в 'web' — значение "
           "'telegram' выставляется только серверным кодом."))
R(row(AREA_CHAT, _SESSION, T_REQ, IN, C_SITE, "player_id", "string", NO, "1048576",
      desc="Идентификатор игрока. Доверяется ТОЛЬКО когда handshake-секрет не задан "
           "(dev); иначе идентичность берётся из подписанного профиля."))
R(row(AREA_CHAT, _SESSION, T_REQ, IN, C_SITE, "signed_context", "string "
      "b64url(payload).b64url(hmac_sha256)", COND,
      "eyJpZCI6…​.QkFTRTY0", desc="Подписанный профиль игрока — единственный "
      "доверенный источник данных в проде. Поля см. строки контракта handshake."))
R(row(AREA_CHAT, _SESSION, T_REQ, IN, C_SITE, "user_context", "object", NO,
      '{"full_name":"Andrey"}',
      desc="Неподписанный профиль. Принимается только в dev (когда секрет handshake "
           "нигде не настроен); иначе обнуляется."))
R(row(AREA_CHAT, _SESSION, T_REQ, IN, C_SITE, "locale", "string (BCP-47)", NO, "es-MX",
      desc="Язык браузера. Определяет стартовый язык ответов и интерфейса."))
R(row(AREA_CHAT, _SESSION, T_REQ, IN, C_SITE, "turnstile_token", "string", NO, "0.ABC…",
      desc="Токен Cloudflare Turnstile. Проверка совещательная: отсутствие токена "
           "или недоступность верификатора пропускаются."))
R(row(AREA_CHAT, _SESSION, T_RESP, OUT, C_US, "session_id", "uuid", YES,
      "3f6d1c2e-…", desc="Идентификатор сессии — ключ всех последующих вызовов."))
R(row(AREA_CHAT, _SESSION, T_RESP, OUT, C_US, "token", "string (JWT HS256)", YES,
      "eyJhbGciOiJIUzI1NiJ9…", desc="Токен сессии: sub=session_id, срок жизни — "
      "настройка продукта (по умолчанию 24 ч). Передаётся как Bearer."))
R(row(AREA_CHAT, _SESSION, T_RESP, OUT, C_US, "topics", "[{slug,title,order}]", YES,
      '[{"slug":"deposits","title":"Депозиты"}]',
      desc="Каталог тем. Тема other всегда последняя."))
R(row(AREA_CHAT, _SESSION, T_RESP, OUT, C_US, "lang", "string (ISO 639-1)", YES, "ru",
      desc="Стартовый язык сессии."))
R(row(AREA_CHAT, _SESSION, T_RESP, OUT, C_US, "languages", "[string]", YES,
      '["en","ru","es"]', desc="Поддерживаемые языки продукта."))

# --- GET /topics, /i18n ----------------------------------------------------
R(row(AREA_CHAT, "GET /api/chat/topics", T_EP, OUT, C_SITE,
      auth=AUTH_NONE, limits="Лимит catalogue:{ip} (600) · Cache-Control 60 с",
      errors="429 rate_limited · 403 bad_widget_key | no_product",
      desc="Каталог тем без создания сессии — для мгновенной отрисовки экрана выбора."))
R(row(AREA_CHAT, "GET /api/chat/topics", T_RESP, OUT, C_US,
      "topics / lang / languages", "[{slug,title,order}] / string / [string]", YES,
      '{"topics":[…],"lang":"ru","languages":["en","ru"]}',
      desc="Каталог тем (other всегда последняя), язык заголовков и список "
           "поддерживаемых языков."))
R(row(AREA_CHAT, "GET /api/chat/topics", T_REQ, IN, C_SITE, "widget_key",
      "query string", COND, "wk_7f3a…", desc="Продукт, чей каталог отдать."))
R(row(AREA_CHAT, "GET /api/chat/topics", T_REQ, IN, C_SITE, "lang | locale",
      "query string", NO, "ru", desc="Язык заголовков тем."))
R(row(AREA_CHAT, "GET /api/chat/i18n", T_EP, OUT, C_SITE,
      auth=AUTH_NONE, limits="Лимит catalogue:{ip} (600) · Cache-Control 60 с",
      errors="429 · 403 bad_widget_key | no_product",
      desc="Строки интерфейса + публичный ключ Turnstile продукта. Позволяет менять "
           "тексты и капчу без редеплоя клиента."))
R(row(AREA_CHAT, "GET /api/chat/i18n", T_RESP, OUT, C_US, "strings",
      "{lang: {key: string}}", YES, '{"ru":{"send":"Отправить"}}',
      desc="Копия интерфейса по языкам."))
R(row(AREA_CHAT, "GET /api/chat/i18n", T_RESP, OUT, C_US, "turnstile_site_key",
      "string|null", NO, "0x4AAA…", desc="Публичный ключ капчи продукта."))
R(row(AREA_CHAT, "GET /api/chat/i18n", T_RESP, OUT, C_US, "languages", "[string]",
      YES, '["en","ru","es"]', desc="Поддерживаемые языки продукта."))

# --- POST /topic -----------------------------------------------------------
R(row(AREA_CHAT, "POST /api/chat/topic", T_EP, IN, C_CLIENT,
      auth=AUTH_SESSION, limits="Общий бюджет chat-op:{ip} (300) на все "
      "сессионные ручки кроме /message",
      errors="401 unauthorized · 404 not_found · 409 session_closed · "
             "400 bad_topic · 429",
      desc="Выбор темы: подгружает базу знаний темы и сдвигает границу истории."))
R(row(AREA_CHAT, "POST /api/chat/topic", T_REQ, IN, C_CLIENT, "session_id", "uuid",
      YES, "3f6d1c2e-…"))
R(row(AREA_CHAT, "POST /api/chat/topic", T_REQ, IN, C_CLIENT, "topic_slug", "string",
      YES, "withdrawals", desc="Slug из каталога тем."))

# --- POST /message ---------------------------------------------------------
_MSG = "POST /api/chat/message"
R(row(AREA_CHAT, _MSG, T_EP, BOTH, C_CLIENT, auth=AUTH_SESSION,
      limits="Порядок гейтов: токен → открытость сессии → лимит по IP → кулдаун → "
             "длина/пустое → малосодержательное (модельный ответ не зовётся, "
             "возвращается 200-подсказка) → инъекция → жёсткий потолок ходов "
             "(2x мягкого капа) → эскалация по ключевым словам → вызов модели",
      errors="401 · 409 session_closed · 429 rate_limited | cooldown · "
             "400 too_long | empty | rejected",
      desc="Один ход диалога. Пишется атомарно: сообщения + счётчик + лог вызова ИИ. "
           "ВАЖНО: модельно-независимые ответы (малосодержательное сообщение, "
           "жёсткий потолок) возвращают только {reply, lang, escalation, "
           "message_count} — остальные ключи ОТСУТСТВУЮТ, клиент обязан "
           "читать их опционально."))
R(row(AREA_CHAT, _MSG, T_REQ, IN, C_CLIENT, "session_id", "uuid", YES, "3f6d1c2e-…"))
R(row(AREA_CHAT, _MSG, T_REQ, IN, C_CLIENT, "text", "string (≤500 симв. по умолч.)",
      YES, "Как вывести деньги?", desc="Сообщение игрока. Предел — настройка продукта."))
R(row(AREA_CHAT, _MSG, T_REQ, IN, C_CLIENT, "closing", "bool", NO, "false",
      desc="Игрок нажал «Проблема решена» — ход трактуется как прощание."))
R(row(AREA_CHAT, _MSG, T_RESP, OUT, C_US, "reply", "string (Markdown-подмножество)",
      YES, "Вывод занимает до 24 часов…",
      desc="Ответ. Поддерживаются **жирный**, *курсив*, `код`, ссылки, списки. "
           "Пустая строка = ход-маршрутизация или эскалация."))
R(row(AREA_CHAT, _MSG, T_RESP, OUT, C_US, "lang", "string (ISO 639-1)", YES, "ru",
      desc="Язык ответа. Может меняться посреди диалога — интерфейс должен следовать."))
R(row(AREA_CHAT, _MSG, T_RESP, OUT, C_US, "message_count", "int", YES, "4",
      desc="Счётчик ходов; на пределе включается передача человеку."))
R(row(AREA_CHAT, _MSG, T_RESP, OUT, C_US, "escalation",
      "{active,final,message,button:{label,url},retention?}", YES,
      '{"active":true,"final":true,…}',
      desc="Блок передачи на человека. Поля описаны отдельными строками ниже."))
R(row(AREA_CHAT, _MSG, T_RESP, OUT, C_US, "suggested_topic", "{slug,title}|null", NO,
      '{"slug":"bonuses","title":"Бонусы"}',
      desc="ОБЯЗАТЕЛЬНАЯ клиентская логика: вызвать POST /topic и переспросить "
           "исходный вопрос. При этом reply пустой — без обработки чат «зависнет»."))
R(row(AREA_CHAT, _MSG, T_RESP, OUT, C_US, "suggestions", "[string] (≤2)", NO,
      '["Какие лимиты на вывод?"]',
      desc="Уточняющие вопросы от лица игрока — кнопки в один тап."))
R(row(AREA_CHAT, _MSG, T_RESP, OUT, C_US, "closing_suggestion", "string|null", NO,
      "Проблема решена.",
      desc="Системная опция завершения: отправить ход с closing=true, затем /resolve."))
R(row(AREA_CHAT, _MSG, T_RESP, OUT, C_US, "resolved", "bool", COND, "true",
      desc="Вопрос выглядит закрытым — показать кнопку «завершить чат». "
           "В модельно-независимых ответах ключ отсутствует."))
R(row(AREA_CHAT, _MSG, T_RESP, OUT, C_US, "cap_notice",
      "{message,escalate_label,finish_label}|null", NO,
      '{"message":"…","escalate_label":"Передать вопрос человеку",'
      '"finish_label":"Завершить чат"}',
      desc="Мягкий лимит сообщений достигнут (или resolved на последних ~2 "
           "ходах перед ним): показать предупреждение и две кнопки — эскалация "
           "(POST /escalate) сверху и завершение (POST /resolve) ниже — вместо "
           "suggestions/resolved. Чат остаётся открытым; на 2x лимита сессия "
           "закрывается принудительно."))

# --- resume / escalate / resolve -------------------------------------------
R(row(AREA_CHAT, "GET /api/chat/session/{id}", T_EP, OUT, C_CLIENT, auth=AUTH_SESSION,
      limits="Бюджет chat-op:{ip} (300)", errors="401 · 404 · 429",
      desc="Возобновление диалога: статус и до 50 последних сообщений."))
R(row(AREA_CHAT, "GET /api/chat/session/{id}", T_RESP, OUT, C_US, "status",
      "'open'|'resolved'|'escalated'", YES, "open",
      desc="Только open допускает новые ходы; иначе 409."))
R(row(AREA_CHAT, "GET /api/chat/session/{id}", T_RESP, OUT, C_US, "history",
      "[{role,content,lang}]", YES, '[{"role":"user","content":"…"}]',
      desc="role ∈ user|assistant."))
R(row(AREA_CHAT, "GET /api/chat/session/{id}", T_RESP, OUT, C_US,
      "session_id / escalated / lang / message_count",
      "uuid / bool / string / int", YES,
      '{"escalated":false,"lang":"ru","message_count":4}',
      desc="Остальные поля возобновления: идентификатор, признак эскалации, "
           "язык сессии и счётчик ходов."))
R(row(AREA_CHAT, "POST /api/chat/topic", T_RESP, OUT, C_US, "ok", "bool", YES,
      '{"ok":true}', desc="Подтверждение выбора темы."))
R(row(AREA_CHAT, "POST /api/chat/escalate", T_EP, BOTH, C_CLIENT, auth=AUTH_SESSION,
      limits="Бюджет chat-op:{ip} (300)", errors="401 · 404 · 409 · 429",
      desc="Явный запрос человека: жёсткая эскалация, сессия закрывается. "
           "Ответ — {escalation: {…}} (контракт escalation ниже)."))
R(row(AREA_CHAT, "POST /api/chat/resolve", T_EP, IN, C_CLIENT, auth=AUTH_SESSION,
      limits="Идемпотентно; не перекрывает уже эскалированную сессию · "
             "бюджет chat-op:{ip} (300)",
      errors="401 · 404 · 429", desc="Игрок завершил чат. "
      'Ответ — {"ok":true,"status":"resolved"}.'))

# --- контракт профиля игрока ----------------------------------------------
_CTX = "Контракт: user_context / signed_context"
_CTX_NOTE = ("Белый список полей, которые видит модель. Всё остальное отбрасывается — "
             "расширение списка требует правки на нашей стороне.")
for key, fmt, example, desc in [
    ("id", "string", "1048576", "Идентификатор игрока в казино."),
    ("full_name", "string", "Андрей Петров",
     "Полное имя; из него берётся имя для приветствия."),
    ("email", "string", "player@example.com", "Почта игрока."),
    ("activation_status", "string", "verified", "Статус активации/верификации."),
    ("country", "string", "ES", "Страна."),
    ("balance", "string", "150.00 USDT", "Баланс строкой, вместе с валютой."),
    ("vip_level", "string", "Gold", "VIP-класс."),
    ("registration_date", "string (ISO-8601)", "2025-03-14", "Дата регистрации."),
]:
    R(row(AREA_CHAT, _CTX, T_REQ, IN, C_SITE, key, fmt, NO, example,
          auth="подпись HMAC-SHA256 секретом продукта",
          desc=desc + " " + _CTX_NOTE if key == "id" else desc))
R(row(AREA_CHAT, _CTX, T_REQ, IN, C_SITE, "iat", "int (unix seconds)", COND,
      "1769606400", auth="часть подписанного payload",
      limits="Возраст токена ограничен 300 с",
      desc="Момент выпуска подписи (антиреплей)."))
R(row(AREA_CHAT, _CTX, T_REQ, IN, C_SITE, "exp", "int (unix seconds)", YES,
      "1769606700", auth="часть подписанного payload",
      errors="401 bad_handshake при истечении",
      desc="Срок годности подписи."))

# --- контракт эскалации ----------------------------------------------------
_ESC = "Контракт: escalation"
R(row(AREA_CHAT, _ESC, T_RESP, OUT, C_US, "active", "bool", YES, "true",
      desc="Есть ли в этом ходе передача на человека."))
R(row(AREA_CHAT, _ESC, T_RESP, OUT, C_US, "final", "bool", YES, "true",
      desc="true — сессия закрыта (дальше 409); false — карточка показана, "
           "диалог продолжается."))
R(row(AREA_CHAT, _ESC, T_RESP, OUT, C_US, "message", "string", COND,
      "Передаю вас специалисту…", desc="Текст карточки на языке ответа."))
R(row(AREA_CHAT, _ESC, T_RESP, OUT, C_US, "button.label", "string", COND,
      "Написать в поддержку", desc="Подпись кнопки."))
R(row(AREA_CHAT, _ESC, T_RESP, OUT, C_US, "button.url", "string (URL)", COND,
      "https://t.me/brand_bot?start=…",
      desc="Куда ведёт кнопка: статическая ссылка контакта или одноразовый "
           "диплинк в Telegram-бота."))
R(row(AREA_CHAT, _ESC, T_RESP, OUT, C_US, "retention", "bool", NO, "true",
      desc="Признак, что передача ведёт в Telegram-бота, а не на статическую ссылку."))

# ===========================================================================
# Partner API — вебхуки казино/CRM
# ===========================================================================
_PU = "POST /partner/{product_id}/player-update"
R(row(AREA_PARTNER, _PU, T_EP, IN, C_CRM, auth=AUTH_PARTNER,
      limits="Частичное обновление: поля со значением null не затираются · "
             "лимит partner:{ip}",
      errors="401 unauthorized (единый ответ на все причины) · "
             "422 invalid_rg_status · 429 rate_limited",
      desc="Push профиля игрока из CRM в наш снапшот. Все вебхуки /partner/* "
           "лимитируются по IP ДО аутентификации."))
R(row(AREA_PARTNER, _PU, T_HDR, IN, C_CRM, "Authorization",
      "Bearer <handshake_secret>", YES, "Bearer 8f2c…",
      auth=AUTH_PARTNER, desc="Тот же секрет продукта, что подписывает handshake."))
R(row(AREA_PARTNER, _PU, T_REQ, IN, C_CRM, "player_id", "string", YES, "1048576",
      desc="Ключ игрока на стороне казино."))
for key, fmt, example in [
    ("full_name", "string", "Андрей Петров"),
    ("email", "string", "player@example.com"),
    ("activation_status", "string", "verified"),
    ("country", "string", "ES"),
    ("balance", "string", "150.00 USDT"),
    ("vip_level", "string", "Gold"),
    ("registration_date", "string (ISO-8601)", "2025-03-14"),
]:
    R(row(AREA_PARTNER, _PU, T_REQ, IN, C_CRM, key, fmt, NO, example,
          desc="Поле профиля. null = не менять."))
for key, desc in [
    ("last_login_at", "Последний вход — сигнал активности."),
    ("last_played_at", "Последняя игра."),
    ("last_deposit_at", "Последний депозит."),
]:
    R(row(AREA_PARTNER, _PU, T_REQ, IN, C_CRM, key, "string (ISO-8601)", NO,
          "2026-07-28T10:15:00Z", desc=desc))
R(row(AREA_PARTNER, _PU, T_RESP, OUT, C_US, "ok / updated", "bool / bool", YES,
      '{"ok":true,"updated":true}', desc="Признак приёма и факта изменения."))
# Orchestrator extension of player-update (all optional, additive).
R(row(AREA_PARTNER, _PU, T_REQ, IN, C_CRM, "rg_status", "enum", NO,
      "self_exclude",
      desc="RG-статус игрока: ok | cool_off | rg_hold | self_exclude. "
           "Источник истины по самоисключению — платформа казино; "
           "self_exclude/rg_hold полностью блокируют проактив и бонусы. "
           "Слать при каждой смене + первичный бэкфилл."))
R(row(AREA_PARTNER, _PU, T_REQ, IN, C_CRM, "rg_status_until",
      "string (ISO-8601)", NO, "2026-10-01T00:00:00Z",
      desc="Срок cool_off (истёкший снимается автоматически)."))
R(row(AREA_PARTNER, _PU, T_REQ, IN, C_CRM, "marketing_consent", "bool", NO,
      "true", desc="Маркетинговое согласие (гейт для Tier-1 рынков)."))
R(row(AREA_PARTNER, _PU, T_REQ, IN, C_CRM, "rg_flags", "object", NO,
      '{"escalating_bets":true}',
      desc="Готовые поведенческие RG-флаги risk-движка казино."))
R(row(AREA_PARTNER, _PU, T_REQ, IN, C_CRM, "timezone", "string", NO,
      "Europe/Istanbul",
      desc="Локальная таймзона игрока (IANA или '+03:00'; geo-IP на стороне "
           "казино) — точность smart-send-time."))
R(row(AREA_PARTNER, _PU, T_REQ, IN, C_CRM,
      "email_opt_in / email_verified / push_opt_in / push_available / "
      "in_app_available / sms_opt_in", "bool", NO, "true",
      desc="Согласия и доступность каналов. Строгий opt-in: канал без "
           "согласия никогда не используется, даже как fallback. "
           "sms_opt_in принимается «на вырост» — sms-канала пока нет."))
R(row(AREA_PARTNER, _PU, T_REQ, IN, C_CRM, "channel_prefs", "object", NO,
      '{"preferred":"email"}',
      desc="Предпочтения игрока по каналам (свободная структура казино); "
           "хранится рядом с согласиями."))

# ---------------------------------------------------------------------------
# Partner API — POST /partner/{product_id}/delivery-status (callback)
# ---------------------------------------------------------------------------
_DS = "POST /partner/{product_id}/delivery-status"
R(row(AREA_PARTNER, _DS, T_EP, IN, C_CRM, auth=AUTH_PARTNER,
      limits="Идемпотентно; статус не движется назад · лимит partner:{ip}",
      errors="401 unauthorized · 422 invalid_status · 429 rate_limited",
      desc="Коллбек прогресса доставки заказа push/in-app, размещённого "
           "оркестратором (см. Исходящие: delivery-endpoint)."))
R(row(AREA_PARTNER, _DS, T_REQ, IN, C_CRM, "delivery_id", "string", YES,
      "dl_a1b2c3d4e5f6", desc="Наш ключ идемпотентности заказа доставки."))
R(row(AREA_PARTNER, _DS, T_REQ, IN, C_CRM, "status", "enum", YES,
      "delivered", desc="queued | sending | sent | delivered | opened | "
      "clicked | bounced. Порядок монотонный: более ранний статус после "
      "более позднего игнорируется."))

# ---------------------------------------------------------------------------
# Outbound — the offer-grant endpoint the CASINO implements
# ---------------------------------------------------------------------------
_OG = "POST <offer_grant_url> (реализует казино)"
R(row(AREA_OUT, _OG, T_EP, OUT, C_US,
      auth="Bearer <partner_out_key продукта>; без ключа заголовок "
           "Authorization не отправляется",
      limits="Идемпотентно по offer_grant_id (повтор возвращает тот же "
             "результат, второй бонус не создаётся) · таймаут "
             "retention.offer_grant_timeout_sec",
      errors="Любой 4xx (кроме 429) = постоянная ошибка, без ретраев; "
             "429 и 5xx — транзиентные, ретраятся",
      desc="Выдача бонуса по его ID из бонус-CMS казино (offer engine). "
           "URL задаётся per product в админке "
           "(PUT /admin/retention/partner-endpoints)."))
R(row(AREA_OUT, _OG, T_REQ, OUT, C_US, "offer_grant_id", "string", YES,
      "og_a1b2c3d4e5f6", desc="Наш детерминированный ключ идемпотентности."))
R(row(AREA_OUT, _OG, T_REQ, OUT, C_US, "player_id / bonus_id / offer_type / "
      "params", "string / string / string / object", YES,
      '{"bonus_id":"BONUS-42"}',
      desc="Кому и какой бонус начислить (bonus_id — ID в бонус-CMS)."))
R(row(AREA_OUT, _OG, T_RESP, IN, C_CRM, "status", "enum", YES, "granted",
      desc="granted | fraud_hold | duplicate | failed (+ partner_ref, "
           "credited_usd, reason). duplicate трактуется как granted; "
           "неизвестный статус — как failed. Персона упоминает бонус "
           "ТОЛЬКО после granted."))

# ---------------------------------------------------------------------------
# Outbound — the delegated delivery endpoint the CASINO implements
# ---------------------------------------------------------------------------
_DL = "POST <delivery_endpoint_url> (реализует казино)"
R(row(AREA_OUT, _DL, T_EP, OUT, C_US,
      auth="Bearer <partner_out_key продукта>; без ключа заголовок "
           "Authorization не отправляется",
      limits="Идемпотентно по delivery_id · таймаут "
             "retention.push_delivery_timeout_sec · транзиентные ошибки "
             "ретраятся (1m/5m/30m), permanent — нет",
      errors="permanent:true в ответе = ретраев не будет",
      desc="Заказ делегированной доставки push/in-app: казино доставляет на "
           "устройство своей инфраструктурой и шлёт прогресс коллбеком "
           "delivery-status."))
R(row(AREA_OUT, _DL, T_REQ, OUT, C_US, "delivery_id / player_id / channel / "
      "title / body / cta_url / ttl_sec",
      "string / string / enum / string / string / string / int", YES,
      '{"channel":"push","ttl_sec":86400}',
      desc="Заказ доставки; channel: push | in_app."))
R(row(AREA_OUT, _DL, T_RESP, IN, C_CRM, "status", "enum", YES, "sent",
      desc="sent | queued | delivered = успех; failed (+ reason, permanent, "
           "provider_ref). Ретраи только при включённой настройке "
           "retention.delivery_retry_enabled."))

# ---------------------------------------------------------------------------
# Outbound — Customer.io transactional email (наш вызов провайдера)
# ---------------------------------------------------------------------------
_EM = "POST https://api[-eu].customer.io/v1/send/email"
R(row(AREA_OUT, _EM, T_EP, OUT, C_US,
      auth="Bearer <email_api_key продукта> (App API key, секрет продукта)",
      limits="Регион us/eu — из конфигурации канала email · HTTP ≥400 = "
             "постоянная ошибка, кроме 429/500/502/503/504 (ретраи 1m/5m/30m)",
      errors="no_email_api_key / no_email_address — постоянные, без ретраев",
      desc="Email-канал мультиканального роутера: транзакционная отправка "
           "через Customer.io. Статусы провайдера прилетают тем же коллбеком "
           "delivery-status."))
R(row(AREA_OUT, _EM, T_REQ, OUT, C_US,
      "to / identifiers.id / subject / body / from? / transactional_message_id?",
      "string / string / string / string (HTML)", YES,
      '{"to":"player@example.com","identifiers":{"id":"1048576"}}',
      desc="from и transactional_message_id — из конфигурации канала email "
           "продукта."))
R(row(AREA_OUT, _EM, T_RESP, IN, C_CRM, "delivery_id", "string", YES,
      "cio_01H…", desc="Референс провайдера; сохраняется как provider_ref "
      "доставки."))

_EV = "POST /partner/{product_id}/event"
R(row(AREA_PARTNER, _EV, T_EP, IN, C_CRM, auth=AUTH_PARTNER,
      limits="Идемпотентно по event_id (доставка at-least-once безопасна) · "
             "пакет до 500 событий · payload до 8 KiB · лимит partner:{ip}",
      errors="401 · 413 batch_too_large · 422 invalid_event · 429",
      desc="Канонический поток событий казино: кормит проактивного агента "
           "и обновляет сигналы активности игрока."))
R(row(AREA_PARTNER, _EV, T_REQ, IN, C_CRM, "event_id", "string (≤64)", YES,
      "evt_2026072801", desc="Ключ идемпотентности. Повтор считается дубликатом, "
      "а не ошибкой."))
R(row(AREA_PARTNER, _EV, T_REQ, IN, C_CRM, "event_name", "string (из таксономии)",
      YES, "deposit_confirmed",
      errors="422 invalid_event на неизвестное имя",
      desc="Имя события. Таксономия — контракт, а не рекомендация; список ниже."))
R(row(AREA_PARTNER, _EV, T_REQ, IN, C_CRM, "player_id | user_id", "string", YES,
      "1048576", desc="Игрок, к которому относится событие."))
R(row(AREA_PARTNER, _EV, T_REQ, IN, C_CRM, "tg_user_id", "int", NO, "584210394",
      desc="Явный получатель, когда один player_id связан с несколькими "
           "Telegram-аккаунтами."))
R(row(AREA_PARTNER, _EV, T_REQ, IN, C_CRM, "timestamp", "string (ISO-8601)", NO,
      "2026-07-28T10:15:00Z",
      desc="Время события (в пакетной/сырой форме принимается и алиас ts). "
           "Значение из будущего приводится к текущему моменту."))
R(row(AREA_PARTNER, _EV, T_REQ, IN, C_CRM, "event_version", "string", NO, "1",
      desc="Версия схемы события на стороне отправителя."))
R(row(AREA_PARTNER, _EV, T_REQ, IN, C_CRM, "payload", "object (≤8 KiB)", NO,
      '{"amount":"50.00","currency":"USDT"}',
      desc="Детали события; может нести и поля профиля — они обновят снапшот."))
R(row(AREA_PARTNER, _EV, T_REQ, IN, C_CRM, "events", "[object] (≤500)", NO,
      '[{"event_id":"…","event_name":"…"}]',
      desc="Пакетная форма. Взаимоисключима с плоскими полями выше."))
R(row(AREA_PARTNER, _EV, T_RESP, OUT, C_US,
      "ok / stored / duplicates / errors", "bool / int / int / [{index,error}]",
      YES, '{"ok":true,"stored":12,"duplicates":3,"errors":[]}',
      desc="Ответ ПАКЕТНОЙ формы: сколько принято, сколько дублей и список "
           "отвергнутых с причинами."))
R(row(AREA_PARTNER, _EV, T_RESP, OUT, C_US, "ok / stored / duplicate / id",
      "bool / bool / bool / int|null", YES,
      '{"ok":true,"stored":true,"duplicate":false,"id":8412}',
      desc="Ответ ОДИНОЧНОЙ (плоской) формы — форма ответа отличается от "
           "пакетной."))

# --- таксономия событий ----------------------------------------------------
_EVENTS = [
    ("session_started", "Сессия", "обновляет last_login_at",
     "Игрок начал сессию на сайте или в приложении."),
    ("session_ended", "Сессия", "обновляет last_login_at", "Игрок завершил сессию."),
    ("deposit_initiated", "Платежи", "взводит таймер брошенной кассы",
     "Депозит начат, ещё не подтверждён. Таймер снимается событием "
     "deposit_confirmed; по истечении может стартовать journey «брошенная касса»."),
    ("deposit_confirmed", "Платежи", "обновляет last_deposit_at",
     "Депозит зачислен — сильнейший позитивный сигнал."),
    ("deposit_failed", "Платежи", "—", "Депозит не прошёл."),
    ("withdrawal_settled", "Платежи", "—", "Вывод завершён."),
    ("bet_settled", "Игра", "обновляет last_played_at",
     "Ставка рассчитана. Сумма проигрыша за 24 ч влияет на тон коммуникации."),
    ("bonus_granted", "Бонусы", "—", "Бонус начислен."),
    ("bonus_claimed", "Бонусы", "—", "Бонус активирован игроком."),
    ("bonus_completed", "Бонусы", "—", "Отыгрыш завершён."),
    ("bonus_expired", "Бонусы", "—", "Бонус сгорел."),
    ("kyc_started", "Верификация", "—", "Верификация начата."),
    ("kyc_approved", "Верификация", "—", "Верификация пройдена."),
    ("kyc_rejected", "Верификация", "—", "Верификация отклонена."),
    ("xp_granted", "Лояльность", "—", "Начислен опыт."),
    ("level_up", "Лояльность", "—", "Новый уровень."),
    ("class_up", "Лояльность", "—", "Новый класс / VIP-тир."),
    ("downgrade", "Лояльность", "—", "Понижение класса."),
    ("highlights_pack_opened", "Активности", "—", "Открыт пак/кейс."),
    ("highlights_pack_completed", "Активности", "—", "Пак завершён."),
    ("check_in_done", "Активности", "—", "Ежедневный чек-ин выполнен."),
    ("mission_completed", "Активности", "—", "Миссия или квест завершён."),
]
for name, group, bridge, desc in _EVENTS:
    R(row(AREA_PARTNER, _EV, T_EVT, IN, C_CRM, name, f"event_name · {group}", "",
          '{"event_name":"%s"}' % name, auth=AUTH_PARTNER,
          limits=bridge, desc=desc))

# ===========================================================================
# Telegram / deeplink
# ===========================================================================
_DPL = "POST /api/retention/deeplink"
R(row(AREA_TG, _DPL, T_EP, BOTH, C_SITE, auth=AUTH_NONE + " + лимит по IP",
      limits="Нонс одноразовый, живёт 120 с (настраивается) · "
             "лимит deeplink:{ip}",
      errors="403 bad_widget_key | retention_disabled · 409 no_bot · "
             "401 bad_handshake · 429",
      desc="Сайт получает одноразовую ссылку в Telegram-бота с профилем игрока."))
R(row(AREA_TG, _DPL, T_REQ, IN, C_SITE, "widget_key", "string (wk_…)", COND, "wk_7f3a…",
      desc="Продукт, чей бот открывать."))
R(row(AREA_TG, _DPL, T_REQ, IN, C_SITE, "signed_context", "string (подпись)", COND,
      "eyJpZCI6…​.QkFTRTY0", desc="Профиль игрока — те же поля и та же подпись, "
      "что в Chat API."))
R(row(AREA_TG, _DPL, T_REQ, IN, C_SITE, "user_context", "object", NO,
      '{"full_name":"Andrey"}',
      desc="Неподписанный профиль — принимается только в dev (как в Chat API); "
           "в проде обнуляется."))
R(row(AREA_TG, _DPL, T_REQ, IN, C_SITE, "escalation", "bool", NO, "true",
      desc="true — точка входа «к менеджеру»."))
R(row(AREA_TG, _DPL, T_REQ, IN, C_SITE, "lang", "string (код или локаль)", NO, "ru",
      desc="Язык, в котором откроется бот."))
R(row(AREA_TG, _DPL, T_RESP, OUT, C_US, "deep_link", "string (URL)", YES,
      "https://t.me/brand_bot?start=Ax7…",
      desc="Ссылка для кнопки. Нонс погашается при первом /start."))
R(row(AREA_TG, _DPL, T_RESP, OUT, C_US, "nonce", "string", YES, "Ax7kQ…",
      desc="Тот же одноразовый токен отдельным полем."))

_WH = "POST /telegram/webhook/{secret}"
R(row(AREA_TG, _WH, T_EP, IN, C_TG,
      auth="путь = webhook-токен продукта (роутинг) + заголовок "
           "X-Telegram-Bot-Api-Secret-Token",
      limits="Всегда быстрый 200 (иначе Telegram ретраит); ход ИИ выполняется в фоне",
      errors="403 bad_secret_token · неизвестный токен пути, неактивный "
             "продукт или выключенный ретеншен → тихий 200",
      desc="Приём апдейтов Telegram — второй фасад над тем же ИИ-ядром."))
R(row(AREA_TG, _WH, T_HDR, IN, C_TG, "X-Telegram-Bot-Api-Secret-Token", "string", YES,
      "b4c9…", desc="Секрет, заданный при регистрации вебхука. Защищает от подделки "
      "апдейтов тем, кто узнал URL."))
R(row(AREA_TG, _WH, T_REQ, IN, C_TG, "update", "object (Telegram Bot API)", YES,
      '{"update_id":1,"message":{…}}', desc="Стандартный формат Telegram."))

# ===========================================================================
# Admin API — интеграция с внешней мастер-панелью
# ===========================================================================
R(row(AREA_ADMIN, "POST /admin/login", T_EP, BOTH, C_ADMIN, auth=AUTH_NONE,
      limits="Лимит попыток входа с IP", errors="401 · 429",
      desc="Вход человека-администратора: email + пароль → JWT."))
R(row(AREA_ADMIN, "POST /admin/login", T_REQ, IN, C_ADMIN, "email / password",
      "string / string", YES, "ops@brand.com",
      desc="Учётная запись администратора."))
R(row(AREA_ADMIN, "POST /admin/login", T_RESP, OUT, C_US,
      "token / role / email / memberships", "string (JWT) / string / string / "
      "[{scope,role,…}]", YES, "eyJhbGciOiJIUzI1NiJ9…",
      desc="Админ-токен (продлевается сам при активной работе) + роль, email "
           "и список членств для проверки прав на своей стороне."))
R(row(AREA_ADMIN, "* /admin/*", T_HDR, IN, C_ADMIN, "Authorization",
      "Bearer <JWT> | Bearer sak_…", YES, "Bearer sak_9f2c…", auth=AUTH_ADMIN,
      limits="Сервисный ключ несёт ОДНУ роль в ОДНОМ скоупе (global/partner/product); "
             "деактивация действует со следующего запроса",
      errors="401 unauthorized · 403 forbidden (вне скоупа)",
      desc="Единая авторизация всего админ-контура. Для машин — ключ sak_…, "
           "который выдаётся человеком-админом и показывается ровно один раз."))
R(row(AREA_ADMIN, "GET /admin/structure", T_EP, OUT, C_ADMIN, auth=AUTH_ADMIN,
      desc="Дерево партнёров и продуктов — точка синхронизации справочников "
           "с внешними системами."))
R(row(AREA_ADMIN, "GET /admin/me", T_EP, OUT, C_ADMIN, auth=AUTH_ADMIN,
      desc="Кто вызывает и какие у него скоупы — для проверки прав на своей стороне."))
R(row(AREA_ADMIN, "POST /admin/partners, PUT /admin/partners/{id}", T_EP, IN, C_ADMIN,
      auth=AUTH_ADMIN, errors="401 · 403", desc="Заведение и правка партнёра."))
R(row(AREA_ADMIN, "POST /admin/products", T_EP, BOTH, C_ADMIN, auth=AUTH_ADMIN,
      limits="Создание продукта автоматически засевает рабочую конфигурацию",
      errors="401 · 403", desc="Заведение казино. В ответе — публичный widget_key "
      "для встраивания виджета."))
R(row(AREA_ADMIN, "PUT /admin/products/{id}", T_REQ, IN, C_ADMIN,
      "name, active, site_url, turnstile_site_key", "string / bool", NO,
      '{"site_url":"https://brand.com"}', auth=AUTH_ADMIN,
      desc="Публичная конфигурация продукта."))
R(row(AREA_ADMIN, "PUT /admin/products/{id}/secrets", T_REQ, IN, C_ADMIN,
      "openai_key_primary, openai_key_fallback, handshake_secret, "
      "telegram_bot_token, player_api_key, turnstile_secret, "
      "partner_out_key, email_api_key", "string", NO,
      '{"handshake_secret":"…"}', auth=AUTH_ADMIN,
      limits="Только запись: наружу возвращаются лишь флаги «задан/не задан»",
      desc="Секреты продукта. Хранятся зашифрованными; через API не читаются. "
           "partner_out_key — Bearer наших исходящих вызовов в казино "
           "(offer-grant, delivery-order); email_api_key — App API key "
           "Customer.io."))
R(row(AREA_ADMIN, "POST /admin/products/{id}/widget-key", T_EP, BOTH, C_ADMIN,
      auth=AUTH_ADMIN, desc="Ротация публичного ключа виджета."))
R(row(AREA_ADMIN, "GET /admin/overview, /timeseries, /by-topic, /by-language",
      T_EP, OUT, C_ADMIN, auth=AUTH_ADMIN,
      limits="Параметры периода from/to, скоуп product_id",
      desc="Метрики поддержки для выгрузки во внешнюю аналитику."))
R(row(AREA_ADMIN, "GET /admin/ai-costs", T_EP, OUT, C_ADMIN, auth=AUTH_ADMIN,
      limits="Разрез по типу вызова: chat | agent | review | media",
      desc="Полные траты на ИИ — сумма сходится со счётом провайдера."))
R(row(AREA_ADMIN, "GET /admin/sessions, /admin/session/{id}", T_EP, OUT, C_ADMIN,
      auth=AUTH_ADMIN, limits="Пагинация page/page_size",
      desc="Журнал диалогов и полная стенограмма с постоимостью — выгрузка "
           "в общее хранилище обращений."))
R(row(AREA_ADMIN, "GET /admin/audit", T_EP, OUT, C_ADMIN, auth=AUTH_ADMIN,
      desc="Аудит административных изменений — для общего SIEM/комплаенс-контура."))
R(row(AREA_ADMIN, "GET /admin/quality/overview, /admin/quality/reviews", T_EP, OUT,
      C_ADMIN, auth=AUTH_ADMIN,
      desc="Оценки качества диалогов (1..5, теги, пробелы в базе знаний)."))
R(row(AREA_ADMIN, "GET /admin/retention/overview, /timeseries, /funnel, "
      "/effectiveness", T_EP, OUT,
      C_ADMIN, auth=AUTH_ADMIN,
      desc="Метрики Telegram-фасада: воронка и результативность касаний."))
R(row(AREA_ADMIN, "PUT /admin/retention/partner-endpoints", T_REQ, IN, C_ADMIN,
      "offer_grant_url, delivery_endpoint_url", "string (https URL)", NO,
      '{"offer_grant_url":"https://cms.brand.com/bonus/grant"}', auth=AUTH_ADMIN,
      limits="Только публичные адреса (защита от SSRF, пиннинг DNS)",
      desc="Куда наш сервис шлёт исходящие вызовы казино: выдача бонуса и "
           "заказ делегированной доставки (см. область «Исходящие»)."))
R(row(AREA_ADMIN, "GET|PUT|POST /admin/integration-checklist*", T_EP, BOTH, C_ADMIN,
      auth=AUTH_ADMIN + "; запись — только глобальный админ",
      desc="Деплой-уровневый чек-лист интеграционных зависимостей от внешних "
           "команд (бонус-CMS, RG-фид, доставка push, Customer.io, BI): "
           "статусы и заметки по каждому пункту."))
R(row(AREA_ADMIN, "GET|PUT /admin/retention/* (оркестратор)", T_EP, BOTH, C_ADMIN,
      auth=AUTH_ADMIN,
      limits="Скоуп продукта; RG-аудит и set-status — глобальный админ",
      desc="Управление оркестратором ретеншена: holdout/uplift, RG-guard, "
           "частотные капы и приоритеты, скоринг/когорты, каталог офферов и "
           "гранты, journeys, шаблоны сценариев, каналы и журнал доставок. "
           "Полный список ручек — /integration-admin."))
R(row(AREA_ADMIN, "POST /admin/api-keys", T_EP, BOTH, C_ADMIN,
      auth="только человек-админ (ключом ключ не выпустить)",
      limits="Открытый текст ключа возвращается РОВНО ОДИН РАЗ; хранится только хеш",
      errors="401 · 403", desc="Выпуск сервисного ключа sak_… для внешней системы."))
R(row(AREA_ADMIN, "GET /admin/mcp/manifest", T_EP, OUT, C_ADMIN, auth=AUTH_ADMIN,
      desc="Машиночитаемое описание админ-инструментов (MCP) для агентских клиентов."))
R(row(AREA_ADMIN, "GET /openapi.json, /docs, /redoc", T_EP, OUT, C_ADMIN,
      auth=AUTH_NONE, limits="Отдаётся только если включено на деплое",
      errors="404, когда выключено",
      desc="Машиночитаемая схема всего API — для генерации клиентов."))

# ===========================================================================
# Исходящие вызовы — что наш сервис требует от чужих систем
# ===========================================================================
_PAPI = "GET <player_api_url> (Player API партнёра)"
R(row(AREA_OUT, _PAPI, T_EP, OUT, C_US,
      auth="Bearer <player_api_key продукта>",
      limits="Таймаут 8 с · срабатывает, когда снапшот старше часа · "
             "только публичные адреса (защита от SSRF, пиннинг IP)",
      errors="Любая ошибка тиха: остаётся прежний снапшот",
      desc="ЛЕНИВОЕ обновление профиля игрока перед ходом бота. Эндпойнт "
           "реализует ПАРТНЁР — это требование к чужой системе."))
R(row(AREA_OUT, _PAPI, T_REQ, OUT, C_US, "player_id", "query string", YES, "1048576",
      desc="Кого запрашиваем."))
R(row(AREA_OUT, _PAPI, T_RESP, IN, C_CRM,
      "full_name, email, activation_status, country, balance, vip_level, "
      "registration_date, last_login_at, last_played_at, last_deposit_at",
      "JSON-объект", NO, '{"full_name":"Андрей","balance":"150.00 USDT"}',
      desc="Ожидаемый ответ партнёра. Лишние поля игнорируются — берётся "
           "тот же белый список, что и везде."))
R(row(AREA_OUT, "Telegram Bot API", T_EP, OUT, C_US,
      auth="токен бота продукта (хранится зашифрованным)",
      limits="Ретраи; недоступный игрок помечается как unreachable",
      desc="sendMessage / sendPhoto / sendVideo / getChatMember / setWebhook."))
R(row(AREA_OUT, "OpenAI API", T_EP, OUT, C_US,
      auth="ключ продукта (1–2 шт.), иначе ключ деплоя",
      limits="Два ключа с автопереключением; предохранитель при сбое провайдера; "
             "таймауты по назначению вызова",
      desc="Генерация ответов и фоновые проходы. Каждый вызов логируется "
           "с атрибуцией расходов."))

# ---------------------------------------------------------------------------
# Оформление
# ---------------------------------------------------------------------------
FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT, size=10)
MONO_FONT = Font(name="Consolas", size=9)
TITLE_FONT = Font(name=FONT, size=15, bold=True, color="1F3864")
NOTE_FONT = Font(name=FONT, size=10, italic=True, color="595959")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
EP_FILL = PatternFill("solid", fgColor="EDF2FB")     # строка-эндпойнт
DIR_FILL = {
    IN: PatternFill("solid", fgColor="E2EFDA"),
    OUT: PatternFill("solid", fgColor="FFF2CC"),
    BOTH: PatternFill("solid", fgColor="FCE4D6"),
}

SUBTITLE = ("Один юнит интеграции — одна строка. Все колонки фильтруются. "
            "Только то, что пересекает границу систем: эндпойнты, поля на проводе, "
            "форматы, авторизация, лимиты и коды ошибок.")


# The style objects the variables sheet reuses, so both sheets of the workbook
# look like one document (passed in rather than imported, so text_variables.py
# never has to import this module back).
def _sheet_style() -> dict:
    return {
        "title": TITLE_FONT, "note": NOTE_FONT, "head_font": HEAD_FONT,
        "head_fill": HEAD_FILL, "body": BODY_FONT, "mono": MONO_FONT,
        "border": BORDER, "group_fill": EP_FILL,
    }


def build_xlsx(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "API"

    ws["A1"] = "Support Chat — API и межмодульное взаимодействие"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 22
    ws["A2"] = SUBTITLE
    ws["A2"].font = NOTE_FONT

    head = 4
    for i, (name, _w) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=head, column=i, value=name)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[head].height = 30

    r = head
    for data in ROWS:
        r += 1
        for i, val in enumerate(data, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.font = MONO_FONT if (i - 1) in MONO_COLS else BODY_FONT
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
            if data[2] == T_EP:
                c.fill = EP_FILL
            if i - 1 == 3 and val in DIR_FILL:
                c.fill = DIR_FILL[val]
                c.alignment = Alignment(vertical="top", horizontal="center",
                                        wrap_text=True)

    for i, (_name, w) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=head + 1, column=1)
    ws.auto_filter.ref = f"A{head}:{get_column_letter(len(COLUMNS))}{r}"

    r += 2
    ws.cell(row=r, column=1,
            value="Файл генерируется из кода: scripts/build_api_reference.py. "
                  "Живая версия — /integration-reference на самом сервисе."
            ).font = NOTE_FONT

    text_variables.add_sheet(wb, _sheet_style())
    wb.save(path)


# ---------------------------------------------------------------------------
# Живая страница — та же таблица с фильтрами и поиском
# ---------------------------------------------------------------------------
PAGE = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Support Chat — API и межмодульное взаимодействие</title>
<style>
  :root {{
    --bg:#0f1320; --panel:#1a2032; --panel2:#232c44; --line:#2e3a57;
    --text:#e7ecf5; --muted:#9aa7c2; --accent:#4f8cff; --good:#36c08a; --warn:#e8b349;
  }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; background:var(--bg); color:var(--text);
    font:14px/1.6 -apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif; }}
  .wrap {{ max-width:1600px; margin:0 auto; padding:28px 20px 70px; }}
  h1 {{ font-size:25px; margin:0 0 6px; }}
  .muted {{ color:var(--muted); }}
  a {{ color:var(--accent); }}
  code {{ background:var(--panel2); border:1px solid var(--line); border-radius:5px;
    padding:1px 6px; font:12.5px/1.5 ui-monospace,Menlo,Consolas,monospace; }}
  .dl {{ display:inline-block; background:rgba(54,192,138,.14); border:1px solid var(--good);
    color:var(--good); border-radius:8px; padding:9px 16px; text-decoration:none;
    font-weight:600; margin:10px 0 4px; }}
  .dl:hover {{ background:rgba(54,192,138,.24); }}
  .filters {{ background:var(--panel); border:1px solid var(--line); border-radius:10px;
    padding:14px 16px; margin:18px 0; display:flex; flex-wrap:wrap; gap:10px 14px;
    align-items:flex-end; position:sticky; top:0; z-index:5; }}
  .filters label {{ display:block; font-size:11px; text-transform:uppercase;
    letter-spacing:.04em; color:var(--muted); margin-bottom:4px; }}
  select, input[type=search] {{ background:var(--panel2); color:var(--text);
    border:1px solid var(--line); border-radius:7px; padding:7px 10px; font-size:13px;
    min-width:150px; }}
  input[type=search] {{ min-width:240px; }}
  .count {{ margin-left:auto; color:var(--muted); font-size:13px; white-space:nowrap; }}
  .reset {{ background:none; border:1px solid var(--line); color:var(--muted);
    border-radius:7px; padding:7px 12px; font-size:13px; cursor:pointer; }}
  .reset:hover {{ color:var(--text); border-color:var(--accent); }}
  .tablewrap {{ overflow-x:auto; border:1px solid var(--line); border-radius:10px; }}
  table {{ width:100%; border-collapse:collapse; font-size:13px; }}
  th, td {{ text-align:left; padding:8px 10px; border-bottom:1px solid var(--line);
    vertical-align:top; }}
  th {{ color:var(--muted); font-size:11px; text-transform:uppercase; letter-spacing:.04em;
    background:var(--panel); position:sticky; top:64px; z-index:2; white-space:nowrap; }}
  tr.ep {{ background:rgba(79,140,255,.07); }}
  tr.ep td {{ font-weight:600; }}
  td.mono {{ font:12.5px/1.5 ui-monospace,Menlo,Consolas,monospace; color:#cbd5e1;
    font-weight:400; }}
  td.dir {{ text-align:center; white-space:nowrap; font-weight:600; }}
  .d-in {{ color:var(--good); }}
  .d-out {{ color:var(--warn); }}
  .d-both {{ color:#e0956b; }}
  .note {{ background:rgba(79,140,255,.08); border:1px solid var(--line);
    border-left:3px solid var(--accent); border-radius:8px; padding:10px 14px; margin:14px 0; }}
  tr.hidden {{ display:none; }}
</style>
</head>
<body>
<div class="wrap">

<h1>API и межмодульное взаимодействие</h1>
<p class="muted">{subtitle}</p>
<p><a class="dl" href="/static/{xlsx}" download>⤓ Скачать эту же таблицу в Excel</a></p>

<div class="note"><b>Направление:</b>
<span class="d-in">Вход</span> — сервис принимает ·
<span class="d-out">Выход</span> — отдаёт ·
<span class="d-both">Вход/Выход</span> — обмен в обе стороны.
Транспорт везде JSON поверх HTTPS. Токен сессии игрока и админский токен — JWT HS256.
Подписанный профиль игрока — <code>b64url(payload).b64url(hmac_sha256)</code>.
Профиль игрока передаётся ОДНИМ и тем же белым списком полей во всех точках обмена.</div>

<div class="filters">{filters}
  <button class="reset" type="button" id="reset">Сбросить</button>
  <span class="count" id="count"></span>
</div>

<div class="tablewrap">
<table id="t">
<thead><tr>{head}</tr></thead>
<tbody>
{body}
</tbody>
</table>
</div>

<p class="muted" style="margin-top:32px">Страница и Excel-файл генерируются из кода одним
скриптом (<code>scripts/build_api_reference.py</code>), поэтому не расходятся между собой.
Здесь только то, что пересекает границу систем; внутренние переменные и тексты продукта
(база знаний, промпт-переменные, копирайт, служебные теги) — на соседней странице
<a href="/integration-variables">переменных и текстов</a> (второй лист того же файла).
Смежные гайды: <a href="/integration">обзор и архитектура</a> ·
<a href="/integration-widget">виджет</a> ·
<a href="/integration-data">данные игрока</a> ·
<a href="/integration-chat-api">Chat API и свой UI</a> ·
<a href="/integration-telegram">retention-бот в Telegram</a> ·
<a href="/integration-admin">интеграция с общей админкой</a>.</p>

</div>
<script>
(function () {{
  var table = document.getElementById('t');
  var rows = Array.prototype.slice.call(table.tBodies[0].rows);
  var selects = Array.prototype.slice.call(document.querySelectorAll('select[data-col]'));
  var search = document.getElementById('q');
  var count = document.getElementById('count');

  function apply() {{
    var term = (search.value || '').trim().toLowerCase();
    var shown = 0;
    rows.forEach(function (tr) {{
      var ok = selects.every(function (sel) {{
        if (!sel.value) return true;
        return tr.cells[+sel.dataset.col].textContent.trim() === sel.value;
      }});
      if (ok && term) ok = tr.textContent.toLowerCase().indexOf(term) !== -1;
      tr.classList.toggle('hidden', !ok);
      if (ok) shown++;
    }});
    count.textContent = shown + ' из ' + rows.length + ' строк';
  }}

  selects.forEach(function (s) {{ s.addEventListener('change', apply); }});
  search.addEventListener('input', apply);
  document.getElementById('reset').addEventListener('click', function () {{
    selects.forEach(function (s) {{ s.value = ''; }});
    search.value = '';
    apply();
  }});
  apply();
}})();
</script>
</body>
</html>
"""


def build_html(path: str, xlsx_name: str) -> None:
    # фильтры
    parts = []
    for col in FILTER_COLS:
        name = COLUMNS[col][0]
        values = sorted({r[col] for r in ROWS if r[col]})
        opts = "".join(f'<option value="{html.escape(v)}">{html.escape(v)}</option>'
                       for v in values)
        parts.append(
            f'<div><label for="f{col}">{html.escape(name)}</label>'
            f'<select id="f{col}" data-col="{col}">'
            f'<option value="">— все —</option>{opts}</select></div>')
    parts.append('<div><label for="q">Поиск по таблице</label>'
                 '<input type="search" id="q" placeholder="поле, путь, код ошибки…">'
                 "</div>")

    head = "".join(f"<th>{html.escape(n)}</th>" for n, _w in COLUMNS)

    body = []
    for data in ROWS:
        cls = ' class="ep"' if data[2] == T_EP else ""
        cells = []
        for i, val in enumerate(data):
            text = html.escape(val or "").replace("\n", "<br>")
            if i == 3:
                dcls = {IN: "d-in", OUT: "d-out", BOTH: "d-both"}.get(val, "")
                cells.append(f'<td class="dir {dcls}">{text}</td>')
            elif i in MONO_COLS:
                cells.append(f'<td class="mono">{text}</td>')
            else:
                cells.append(f"<td>{text}</td>")
        body.append(f"<tr{cls}>" + "".join(cells) + "</tr>")

    with open(path, "w", encoding="utf-8") as fh:
        fh.write(PAGE.format(subtitle=html.escape(SUBTITLE), xlsx=xlsx_name,
                             filters="\n".join(parts), head=head,
                             body="\n".join(body)))


if __name__ == "__main__":
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    xlsx_name = "api-reference.xlsx"
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        root, "frontend", xlsx_name)
    html_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        root, "frontend", "integration-reference.html")
    vars_path = os.path.join(root, "frontend", "integration-variables.html")
    build_xlsx(xlsx_path)
    build_html(html_path, os.path.basename(xlsx_path))
    text_variables.build_html(vars_path, os.path.basename(xlsx_path))
    print(f"rows: {len(ROWS)} api + {len(text_variables.ROWS)} variables\n"
          f"written: {xlsx_path}\nwritten: {html_path}\nwritten: {vars_path}")
